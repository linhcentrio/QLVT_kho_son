#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MODULE XUẤT BÁO CÁO - EXPORT_REPORT.PY
Excel 4 sheet, CSV, Text report
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import json

class ReportExporter:
    """Xuất báo cáo từ DataFrame phiếu"""

    def __init__(self, df_phieu):
        """Khởi tạo với DataFrame phiếu"""
        self.df_phieu = df_phieu
        self._tinh_thong_ke()

    def _tinh_thong_ke(self):
        """Tính toán thống kê"""
        if self.df_phieu is None or len(self.df_phieu) == 0:
            self.tong_vat_tu = 0
            self.tong_vat_tu_30 = 0
            self.tong_doanh_thu = 0
            self.so_phieu = 0
            self.so_dong = 0
            return

        self.tong_vat_tu = self.df_phieu['Thành tiền'].sum()
        self.tong_vat_tu_30 = self.df_phieu['Vật tư 30%'].sum()
        self.tong_doanh_thu = self.df_phieu['Doanh thu BG'].sum()
        self.so_phieu = self.df_phieu['Số BG'].nunique()
        self.so_dong = len(self.df_phieu)

    def _them_sheet_chi_tiet(self, wb):
        """Thêm sheet chi tiết tất cả dòng phiếu"""
        ws = wb.active
        ws.title = "Chi tiết"

        # Header
        headers = self.df_phieu.columns.tolist()
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data
        for row_num, row_data in enumerate(self.df_phieu.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _them_sheet_tom_tat(self, wb):
        """Thêm sheet tóm tắt chỉ số tổng quát"""
        ws = wb.create_sheet("Tóm tắt")

        # Header
        ws['A1'] = "TỔNG HỢP CHỈ SỐ"
        ws['A1'].font = Font(bold=True, size=14)

        # Data
        data = [
            ['Tổng vật tư xuất', f"{self.tong_vat_tu:,.0f} đ"],
            ['Tổng vật tư 30%', f"{self.tong_vat_tu_30:,.0f} đ"],
            ['Tổng doanh thu BG', f"{self.tong_doanh_thu:,.0f} đ"],
            ['Số phiếu BG', self.so_phieu],
            ['Số dòng vật tư', self.so_dong],
        ]

        for row_num, row_data in enumerate(data, 3):
            ws[f'A{row_num}'] = row_data[0]
            ws[f'B{row_num}'] = row_data[1]
            ws[f'A{row_num}'].font = Font(bold=True)

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20

    def _them_sheet_theo_to(self, wb):
        """Thêm sheet thống kê theo tổ"""
        ws = wb.create_sheet("Theo Tổ")

        if self.df_phieu is None or len(self.df_phieu) == 0:
            return

        # Tính thống kê
        stats = self.df_phieu.groupby('Tổ').agg({
            'Thành tiền': 'sum',
            'STT': 'count'
        }).reset_index()

        # Header
        ws['A1'] = 'Tổ'
        ws['B1'] = 'Tổng tiền (đ)'
        ws['C1'] = 'Số dòng'

        for cell in [ws['A1'], ws['B1'], ws['C1']]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        # Data
        for row_num, row_data in enumerate(stats.values, 2):
            ws[f'A{row_num}'] = row_data[0]
            ws[f'B{row_num}'] = row_data[1]
            ws[f'C{row_num}'] = row_data[2]

        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15

    def _them_sheet_theo_mat_hang(self, wb):
        """Thêm sheet thống kê theo mặt hàng"""
        ws = wb.create_sheet("Theo Mặt hàng")

        if self.df_phieu is None or len(self.df_phieu) == 0:
            return

        # Tính thống kê
        stats = self.df_phieu.groupby(['Mã hàng', 'Tên hàng']).agg({
            'Số lượng': 'sum',
            'Thành tiền': 'sum'
        }).reset_index()

        # Header
        ws['A1'] = 'Mã'
        ws['B1'] = 'Tên hàng'
        ws['C1'] = 'Tổng SL'
        ws['D1'] = 'Tổng tiền (đ)'

        for cell in [ws['A1'], ws['B1'], ws['C1'], ws['D1']]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        # Data
        for row_num, row_data in enumerate(stats.values, 2):
            ws[f'A{row_num}'] = row_data[0]
            ws[f'B{row_num}'] = row_data[1]
            ws[f'C{row_num}'] = row_data[2]
            ws[f'D{row_num}'] = row_data[3]

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20

    def xuat_phieu_excel(self, file_path):
        """Xuất Excel 4 sheet: Chi tiết, Tóm tắt, Theo tổ, Theo mặt hàng"""
        wb = Workbook()

        # Sheet 1: Chi tiết
        self._them_sheet_chi_tiet(wb)

        # Sheet 2-4: Thống kê
        self._them_sheet_tom_tat(wb)
        self._them_sheet_theo_to(wb)
        self._them_sheet_theo_mat_hang(wb)

        # Lưu file
        wb.save(file_path)
        print(f"✅ Xuất Excel thành công: {file_path}")

    def xuat_csv(self, file_path):
        """Xuất CSV format"""
        self.df_phieu.to_csv(file_path, index=False, encoding='utf-8')
        print(f"✅ Xuất CSV thành công: {file_path}")

    def xuat_text_report(self):
        """Xuất báo cáo dạng text"""
        report = f"""
╔════════════════════════════════════════════════════════════════════╗
║          BÁO CÁO XUẤT VẬT TƯ - {datetime.now().strftime('%d/%m/%Y')}          ║
╚════════════════════════════════════════════════════════════════════╝

TỔNG HỢP CHỈ SỐ:
─────────────────────────────────────────────────────────────────────
Tổng vật tư xuất     : {self.tong_vat_tu:>20,.0f} đ
Tổng vật tư 30%      : {self.tong_vat_tu_30:>20,.0f} đ
Tổng doanh thu BG    : {self.tong_doanh_thu:>20,.0f} đ
Số phiếu BG          : {self.so_phieu:>20}
Số dòng vật tư       : {self.so_dong:>20}
─────────────────────────────────────────────────────────────────────

Báo cáo được tạo lúc: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
        """
        return report


# ═══════════════════════════════════════════════════════════════════════════
# TEST
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # Test data
    test_data = {
        'STT': [1, 2],
        'CVDV': ['Sơn', 'Sơn'],
        'Ngày tháng': ['01/01/2026', '01/01/2026'],
        'Số BG': ['BG001', 'BG001'],
        'BKS': ['37C12345', '37C12345'],
        'Hiệu xe': ['VF3 Fadil', 'VF3 Fadil'],
        'Màu': ['Đen', 'Xanh'],
        'Khách hàng': ['Khách 1', 'Khách 1'],
        'Tên hàng': ['Màu đen', 'Màu xanh'],
        'Mã hàng': ['GB0', 'GUN'],
        'Số lượng': [100, 150],
        'Đơn giá': [734, 655],
        'Thành tiền': [73400, 98250],
        'Vật tư 30%': [1200000, 1350000],
        'Doanh thu BG': [4000000, 4500000],
        'KTV': ['KTV1', 'KTV2'],
        'Tổ': ['Tổ sơn', 'Tổ sơn'],
        'Ghi chú': ['', 'Ghi chú test']
    }

    df_test = pd.DataFrame(test_data)
    exporter = ReportExporter(df_test)

    print(exporter.xuat_text_report())
    exporter.xuat_phieu_excel('test_report.xlsx')
    exporter.xuat_csv('test_report.csv')
